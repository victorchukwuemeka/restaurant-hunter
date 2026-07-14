"""
clean_leads.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Cleans and organizes leads_raw.csv into a neat, readable
table — sorted, deduped, categorized, and ready to pitch.

Now handles multi-city, multi-category data with
per-category WhatsApp message templates.

INSTALL:
  pip install pandas tabulate openpyxl

USAGE:
  python clean_leads.py

OUTPUT:
  leads_clean.csv    — clean CSV
  leads_clean.xlsx   — Excel with formatting (open this)
  leads_preview.txt  — pretty table preview in terminal
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import pandas as pd
from tabulate import tabulate
from colorama import Fore, Style, init

init(autoreset=True)

INPUT_CSV   = "leads_raw.csv"
OUTPUT_CSV  = "leads_clean.csv"
OUTPUT_XLSX = "leads_clean.xlsx"
OUTPUT_TXT  = "leads_preview.txt"

DEMO_LINK   = "https://labella-lagos.netlify.app"  # ← your live site

# ── NICHE-SPECIFIC WHATSAPP TEMPLATES ─────────────────────────
# Structure: { category: { type_keyword: message } }
# The first matching keyword in the business's OSM type wins.
# Falls back to the category default, then the global default.

NICHE_TEMPLATES = {
    "Food & Dining": {
        "restaurant": (
            "Hi {name} 👋\n\n"
            "I noticed your restaurant doesn't have a website yet.\n\n"
            "I built this demo for a Lagos restaurant:\n"
            "👉 {demo_link}\n\n"
            "I can build something like this for you — with your menu, "
            "food photos, table reservations, and delivery info.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "cafe": (
            "Hi {name} 👋\n\n"
            "I noticed your cafe doesn't have a website yet.\n\n"
            "A website can showcase your menu, ambiance, and "
            "specialty drinks — people search 'cafes near me' every day.\n\n"
            "I build clean websites for cafes — with your menu, "
            "photo gallery, WiFi info, and location map.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "fast_food": (
            "Hi {name} 👋\n\n"
            "I noticed your spot doesn't have a website yet.\n\n"
            "People search online for quick food options near them.\n\n"
            "I build websites for fast food spots — with your full menu, "
            "price list, delivery WhatsApp link, and Google Maps pin.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "food_court": (
            "Hi {name} 👋\n\n"
            "I noticed your food court doesn't have a website yet.\n\n"
            "A website can list all your vendors, menus, and "
            "operating hours — making it easy for customers to find you.\n\n"
            "I build websites for food courts — with vendor listings, "
            "menus, photo galleries, and event announcements.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "bakery": (
            "Hi {name} 👋\n\n"
            "I noticed your bakery doesn't have a website yet.\n\n"
            "A website with your cake gallery and price list can "
            "bring in custom orders — especially for weddings and events.\n\n"
            "I build beautiful websites for bakeries — with product "
            "galleries, custom order forms, and WhatsApp ordering.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your food business doesn't have a website yet.\n\n"
            "I built this demo for a Lagos restaurant:\n"
            "👉 {demo_link}\n\n"
            "I can build something like this for you — with your menu, "
            "your photos, and an online ordering link.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
    },

    "Beauty & Salons": {
        "hairdresser": (
            "Hi {name} 👋\n\n"
            "I noticed your salon doesn't have a website yet.\n\n"
            "Clients Google 'hair salon near me' before booking. "
            "A website with your best styles can bring them straight to you.\n\n"
            "I build websites for hair salons — with a portfolio gallery, "
            "service menu + prices, online booking, and location.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "beauty": (
            "Hi {name} 👋\n\n"
            "I noticed your beauty shop doesn't have a website yet.\n\n"
            "A website can showcase your products, treatments, and "
            "before/after results — and let clients book online.\n\n"
            "I build websites for beauty businesses — with product catalogs, "
            "treatment menus, gallery, and WhatsApp booking.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "cosmetics": (
            "Hi {name} 👋\n\n"
            "I noticed your cosmetics shop doesn't have a website yet.\n\n"
            "Customers search for beauty products online before buying. "
            "A website can display your full range and drive foot traffic.\n\n"
            "I build websites for cosmetics shops — with product showcases, "
            "price lists, brand pages, and location info.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "perfume": (
            "Hi {name} 👋\n\n"
            "I noticed your perfume shop doesn't have a website yet.\n\n"
            "A website can showcase your fragrance collection and "
            "help customers discover new scents before visiting.\n\n"
            "I build websites for perfume shops — with product catalogs, "
            "brand collections, price lists, and store location.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "beauty_salon": (
            "Hi {name} 👋\n\n"
            "I noticed your salon doesn't have a website yet.\n\n"
            "Clients look for beauty treatments online — a website "
            "with your services and prices builds instant trust.\n\n"
            "I build websites for beauty salons — with treatment menus, "
            "before/after gallery, online booking, and WhatsApp contact.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your salon doesn't have a website yet.\n\n"
            "A website can help you get more walk-in clients and "
            "show off your best work online.\n\n"
            "I build professional websites for salons — with online "
            "booking, a portfolio gallery, and your full service menu.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
    },

    "Health & Medical": {
        "pharmacy": (
            "Hi {name} 👋\n\n"
            "I noticed your pharmacy doesn't have a website yet.\n\n"
            "People search for pharmacies and drug prices online. "
            "A website with your location, hours, and services "
            "helps patients find you faster.\n\n"
            "I build websites for pharmacies — with service list, "
            "operating hours, delivery info, and Google Maps location.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "clinic": (
            "Hi {name} 👋\n\n"
            "I noticed your clinic doesn't have a website yet.\n\n"
            "Patients check online before choosing a clinic. "
            "A professional website builds trust and fills appointment slots.\n\n"
            "I build websites for clinics — with doctor profiles, "
            "services, appointment booking, and contact details.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "hospital": (
            "Hi {name} 👋\n\n"
            "I noticed your hospital doesn't have a website yet.\n\n"
            "A website makes your departments, doctors, and services "
            "easily discoverable — and gives patients confidence.\n\n"
            "I build websites for hospitals — with department pages, "
            "doctor directories, appointment forms, and emergency info.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "dentist": (
            "Hi {name} 👋\n\n"
            "I noticed your dental practice doesn't have a website yet.\n\n"
            "People search 'dentist near me' before booking. "
            "A website with your services and before/after photos "
            "builds instant trust.\n\n"
            "I build websites for dental practices — with treatment menus, "
            "gallery, appointment booking, and patient testimonials.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "doctors": (
            "Hi {name} 👋\n\n"
            "I noticed your practice doesn't have a website yet.\n\n"
            "Patients research doctors online before booking. "
            "A website with your qualifications, specializations, "
            "and booking form can bring in new patients.\n\n"
            "I build websites for medical practices — with doctor bios, "
            "services, appointment scheduling, and contact info.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "optician": (
            "Hi {name} 👋\n\n"
            "I noticed your eye clinic doesn't have a website yet.\n\n"
            "People search for eye tests and glasses shops online.\n\n"
            "I build websites for opticians — with frame galleries, "
            "service menus, price ranges, and booking forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "laboratory": (
            "Hi {name} 👋\n\n"
            "I noticed your lab doesn't have a website yet.\n\n"
            "A website can list your test types, prices, and sample "
            "collection hours — patients and hospitals search for this.\n\n"
            "I build websites for medical labs — with test catalogs, "
            "price lists, results inquiry forms, and location info.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your health practice doesn't have a website yet.\n\n"
            "Patients often search online before choosing a doctor "
            "or clinic.\n\n"
            "I build professional websites for health businesses — "
            "with appointment booking, service lists, and contact info.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
    },

    "Auto & Garage": {
        "car_repair": (
            "Hi {name} 👋\n\n"
            "I noticed your garage doesn't have a website yet.\n\n"
            "Car owners Google 'mechanic near me' when their car breaks down. "
            "A website with your services and location brings them to you.\n\n"
            "I build websites for auto repair shops — with service lists, "
            "price estimates, garage photos, and WhatsApp booking.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "car_parts": (
            "Hi {name} 👋\n\n"
            "I noticed your parts shop doesn't have a website yet.\n\n"
            "Mechanics and car owners search for spare parts online.\n\n"
            "I build websites for auto parts shops — with product catalogs, "
            "brand listings, price ranges, and contact info.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "tyres": (
            "Hi {name} 👋\n\n"
            "I noticed your tyre shop doesn't have a website yet.\n\n"
            "Drivers search for tyre prices and brands before buying.\n\n"
            "I build websites for tyre shops — with brand pages, size "
            "guides, price lists, and booking for fitting.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "car": (
            "Hi {name} 👋\n\n"
            "I noticed your car dealership doesn't have a website yet.\n\n"
            "Buyers research cars online before visiting a dealer.\n\n"
            "I build websites for car dealers — with vehicle galleries, "
            "specs, price lists, financing info, and WhatsApp contact.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "car_wash": (
            "Hi {name} 👋\n\n"
            "I noticed your car wash doesn't have a website yet.\n\n"
            "Drivers search 'car wash near me' regularly. "
            "A website with your packages, prices, and location "
            "makes you the easy choice.\n\n"
            "I build websites for car washes — with service packages, "
            "price cards, location map, and booking link.\n\n"
            "Ready in 1 week. Interested?"
        ),
        "car_cleaning": (
            "Hi {name} 👋\n\n"
            "I noticed your car wash doesn't have a website yet.\n\n"
            "Drivers search 'car wash near me' regularly. "
            "A website with your packages, prices, and location "
            "makes you the easy choice.\n\n"
            "I build websites for car washes — with service packages, "
            "price cards, location map, and booking link.\n\n"
            "Ready in 1 week. Interested?"
        ),
        "bicycle_repair": (
            "Hi {name} 👋\n\n"
            "I noticed your bike shop doesn't have a website yet.\n\n"
            "Cyclists search online for repair shops and parts.\n\n"
            "I build websites for bike shops — with service menus, "
            "parts catalog, brand info, and booking forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your auto shop doesn't have a website yet.\n\n"
            "Car owners Google mechanics and car washes near them "
            "every day.\n\n"
            "I build websites for auto businesses — with your services, "
            "price list, location map, and contact info.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
    },

    "Fashion & Retail": {
        "clothes": (
            "Hi {name} 👋\n\n"
            "I noticed your fashion store doesn't have a website yet.\n\n"
            "Shoppers browse fashion online before visiting. "
            "A website with your latest collection can drive "
            "foot traffic and WhatsApp orders.\n\n"
            "I build websites for clothing stores — with lookbook galleries, "
            "size guides, new arrivals page, and WhatsApp ordering.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "fashion": (
            "Hi {name} 👋\n\n"
            "I noticed your fashion brand doesn't have a website yet.\n\n"
            "A website gives your brand a real presence — beyond "
            "Instagram. Customers can browse your full collection.\n\n"
            "I build websites for fashion brands — with lookbook, "
            "new arrivals, styling tips blog, and WhatsApp orders.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "shoes": (
            "Hi {name} 👋\n\n"
            "I noticed your shoe store doesn't have a website yet.\n\n"
            "People search for shoe brands and styles online.\n\n"
            "I build websites for shoe stores — with product galleries, "
            "size charts, brand pages, and contact info.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "bag": (
            "Hi {name} 👋\n\n"
            "I noticed your bags shop doesn't have a website yet.\n\n"
            "A website can showcase your bags collection and let "
            "customers inquire or order via WhatsApp.\n\n"
            "I build websites for bags and accessories shops — with "
            "product catalogs, price lists, and ordering.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "jewelry": (
            "Hi {name} 👋\n\n"
            "I noticed your jewelry store doesn't have a website yet.\n\n"
            "A website with high-quality photos of your pieces can "
            "attract high-value customers who search online.\n\n"
            "I build elegant websites for jewelry stores — with product "
            "galleries, price ranges, custom order forms, and store info.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "watches": (
            "Hi {name} 👋\n\n"
            "I noticed your watch shop doesn't have a website yet.\n\n"
            "Watch buyers research brands and prices online first.\n\n"
            "I build websites for watch shops — with brand pages, "
            "product galleries, service info, and contact details.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "accessories": (
            "Hi {name} 👋\n\n"
            "I noticed your accessories shop doesn't have a website yet.\n\n"
            "A website can display your full range and let customers "
            "browse before visiting your store.\n\n"
            "I build websites for accessories shops — with product "
            "catalogs, style galleries, and WhatsApp ordering.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your store doesn't have a website yet.\n\n"
            "A website can showcase your collection and let customers "
            "shop or inquire online.\n\n"
            "I build clean, modern websites for fashion businesses — "
            "with product galleries, WhatsApp ordering, and more.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
    },

    "Fitness & Gym": {
        "fitness_centre": (
            "Hi {name} 👋\n\n"
            "I noticed your gym doesn't have a website yet.\n\n"
            "People search 'gyms near me' before signing up. "
            "A website with your classes, trainers, and pricing "
            "makes you the obvious choice.\n\n"
            "I build websites for gyms — with class schedules, trainer "
            "profiles, membership plans, and trial sign-up forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "sports_centre": (
            "Hi {name} 👋\n\n"
            "I noticed your sports center doesn't have a website yet.\n\n"
            "A website can list your facilities, booking options, "
            "and membership plans — and make you easy to find.\n\n"
            "I build websites for sports centers — with facility galleries, "
            "booking forms, event calendars, and pricing.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your fitness business doesn't have a website yet.\n\n"
            "People search online before joining a gym.\n\n"
            "I build websites for fitness businesses — with class schedules, "
            "membership plans, and online sign-up.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
    },

    "Education & Training": {
        "school": (
            "Hi {name} 👋\n\n"
            "I noticed your school doesn't have a website yet.\n\n"
            "Parents search online when choosing schools for their children. "
            "A website with your programs, results, and facilities "
            "builds instant trust.\n\n"
            "I build websites for schools — with program pages, "
            "admission forms, photo galleries, and parent portals.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "college": (
            "Hi {name} 👋\n\n"
            "I noticed your college doesn't have a website yet.\n\n"
            "Students and parents research institutions online before "
            "applying. A professional website makes a strong first impression.\n\n"
            "I build websites for colleges — with course catalogs, "
            "admission info, campus galleries, and inquiry forms.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "university": (
            "Hi {name} 👋\n\n"
            "I noticed your institution doesn't have a website yet.\n\n"
            "A website is essential for attracting students and "
            "establishing credibility.\n\n"
            "I build websites for educational institutions — with faculty "
            "pages, course listings, campus tours, and admission forms.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "kindergarten": (
            "Hi {name} 👋\n\n"
            "I noticed your school doesn't have a website yet.\n\n"
            "Parents search for nurseries and kindergartens online. "
            "A website with your curriculum, photos, and fees "
            "helps parents feel confident choosing you.\n\n"
            "I build websites for nurseries — with program info, "
            "photo galleries, fee structures, and enrollment forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "language_school": (
            "Hi {name} 👋\n\n"
            "I noticed your language school doesn't have a website yet.\n\n"
            "Students search for language courses and class schedules online.\n\n"
            "I build websites for language schools — with course listings, "
            "level guides, schedule calendars, and enrollment forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your school/training center doesn't have a website yet.\n\n"
            "Parents and students search online before choosing a school.\n\n"
            "I build professional websites for schools — with program info, "
            "admission forms, photo galleries, and contact details.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
    },

    "Events & Hospitality": {
        "hotel": (
            "Hi {name} 👋\n\n"
            "I noticed your hotel doesn't have a website yet.\n\n"
            "Travelers search for hotels online before booking. "
            "A website with room photos, pricing, and direct "
            "booking can cut your OTA commissions.\n\n"
            "I build websites for hotels — with room galleries, "
            "rate cards, booking forms, and virtual tours.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "hostel": (
            "Hi {name} 👋\n\n"
            "I noticed your hostel doesn't have a website yet.\n\n"
            "Budget travelers and backpackers search for hostels online.\n\n"
            "I build websites for hostels — with dorm/bunk photos, "
            "amenities list, pricing, and booking forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "guest_house": (
            "Hi {name} 👋\n\n"
            "I noticed your guest house doesn't have a website yet.\n\n"
            "Guests search online for affordable lodging with good reviews.\n\n"
            "I build websites for guest houses — with room galleries, "
            "amenities, pricing, reviews display, and booking form.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "bar": (
            "Hi {name} 👋\n\n"
            "I noticed your bar doesn't have a website yet.\n\n"
            "People search 'bars near me' and 'happy hour deals' online. "
            "A website with your menu, events, and vibe photos "
            "pulls in the crowd.\n\n"
            "I build websites for bars — with drink menus, event calendars, "
            "photo galleries, and location map.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "nightclub": (
            "Hi {name} 👋\n\n"
            "I noticed your club doesn't have a website yet.\n\n"
            "Partygoers search for clubs and events online.\n\n"
            "I build websites for nightclubs — with event lineups, "
            "photo/video galleries, VIP booking, and guest list forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "pub": (
            "Hi {name} 👋\n\n"
            "I noticed your pub doesn't have a website yet.\n\n"
            "A website can showcase your atmosphere, menu, and "
            "weekly events — and help people find you.\n\n"
            "I build websites for pubs — with food/drink menus, "
            "event schedules, photo galleries, and location info.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "event_venue": (
            "Hi {name} 👋\n\n"
            "I noticed your event venue doesn't have a website yet.\n\n"
            "Event planners and hosts search for venues online. "
            "A website with capacity info, photos, and availability "
            "can fill your calendar.\n\n"
            "I build websites for event venues — with photo galleries, "
            "capacity charts, pricing, availability calendar, and inquiry forms.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your venue doesn't have a website yet.\n\n"
            "Guests and event hosts search online first.\n\n"
            "I build websites for hotels and event spaces — with room/event "
            "galleries, booking forms, pricing, and location maps.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
    },

    "Professional Services": {
        "lawyer": (
            "Hi {name} 👋\n\n"
            "I noticed your law firm doesn't have a website yet.\n\n"
            "Clients search for lawyers and legal services online. "
            "A professional website builds the trust needed to "
            "win that first consultation.\n\n"
            "I build websites for law firms — with practice area pages, "
            "lawyer profiles, case results, and consultation booking.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "accountant": (
            "Hi {name} 👋\n\n"
            "I noticed your firm doesn't have a website yet.\n\n"
            "Businesses search for accountants and tax advisors online.\n\n"
            "I build websites for accounting firms — with service pages, "
            "team bios, client testimonials, and contact forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "insurance": (
            "Hi {name} 👋\n\n"
            "I noticed your insurance office doesn't have a website yet.\n\n"
            "People compare insurance providers online before buying.\n\n"
            "I build websites for insurance businesses — with product pages, "
            "quote request forms, claims info, and agent profiles.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "estate_agent": (
            "Hi {name} 👋\n\n"
            "I noticed your real estate agency doesn't have a website yet.\n\n"
            "Property buyers and renters search online first. "
            "A website with listings can generate leads 24/7.\n\n"
            "I build websites for real estate agencies — with property "
            "listings, photo galleries, search filters, and inquiry forms.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "bank": (
            "Hi {name} 👋\n\n"
            "I noticed your branch doesn't have a local web presence yet.\n\n"
            "A website with your services, hours, and contact info "
            "helps customers find and choose your branch.\n\n"
            "I build websites for bank branches — with service menus, "
            "ATM locator, hours, and contact forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "it": (
            "Hi {name} 👋\n\n"
            "I noticed your IT company doesn't have a website yet.\n\n"
            "A website is your digital storefront — clients expect "
            "tech companies to have a strong online presence.\n\n"
            "I build websites for IT companies — with service pages, "
            "portfolio/case studies, team bios, and contact forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your business doesn't have a website yet.\n\n"
            "A professional website builds trust with clients and makes "
            "your business look credible.\n\n"
            "I build clean, modern websites for professional firms — "
            "with service pages, team bios, and contact forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
    },

    "Home Services": {
        "furniture": (
            "Hi {name} 👋\n\n"
            "I noticed your furniture store doesn't have a website yet.\n\n"
            "Customers browse furniture online before visiting showrooms. "
            "A website with your catalog can bring them in.\n\n"
            "I build websites for furniture stores — with product catalogs, "
            "room galleries, price ranges, and custom order forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "hardware": (
            "Hi {name} 👋\n\n"
            "I noticed your hardware store doesn't have a website yet.\n\n"
            "Builders and contractors search for materials and prices online.\n\n"
            "I build websites for hardware stores — with product catalogs, "
            "price lists, bulk order forms, and delivery info.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "electrical": (
            "Hi {name} 👋\n\n"
            "I noticed your electrical shop doesn't have a website yet.\n\n"
            "Customers search for electrical fittings and brands online.\n\n"
            "I build websites for electrical shops — with product pages, "
            "brand listings, price ranges, and contact info.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "paint": (
            "Hi {name} 👋\n\n"
            "I noticed your paint shop doesn't have a website yet.\n\n"
            "Homeowners and contractors search for paint colors and brands.\n\n"
            "I build websites for paint shops — with color catalogs, "
            "brand pages, price lists, and consultation booking.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "tiles": (
            "Hi {name} 👋\n\n"
            "I noticed your tiles shop doesn't have a website yet.\n\n"
            "Builders and homeowners browse tiles online before buying.\n\n"
            "I build websites for tiles shops — with product galleries, "
            "style collections, price ranges, and showroom info.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your business doesn't have a website yet.\n\n"
            "Homeowners Google service providers every day.\n\n"
            "I build websites for home service businesses — with your "
            "portfolio, service list, price ranges, and WhatsApp contact.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
    },

    "Electronics & Tech": {
        "electronics": (
            "Hi {name} 👋\n\n"
            "I noticed your electronics shop doesn't have a website yet.\n\n"
            "Customers compare electronics prices and specs online.\n\n"
            "I build websites for electronics shops — with product catalogs, "
            "brand pages, price comparisons, and WhatsApp ordering.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "computer": (
            "Hi {name} 👋\n\n"
            "I noticed your computer shop doesn't have a website yet.\n\n"
            "People search for computers, laptops, and accessories online.\n\n"
            "I build websites for computer shops — with product listings, "
            "specs, price lists, and service/repair booking.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "mobile_phone": (
            "Hi {name} 👋\n\n"
            "I noticed your phone shop doesn't have a website yet.\n\n"
            "Customers browse phone prices and deals online first.\n\n"
            "I build websites for phone shops — with product catalogs, "
            "brand pages, price lists, and trade-in inquiry forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "telecom": (
            "Hi {name} 👋\n\n"
            "I noticed your telecom business doesn't have a website yet.\n\n"
            "A website can display your services, data plans, "
            "and coverage info — helping customers choose you.\n\n"
            "I build websites for telecom businesses — with service pages, "
            "plan comparisons, coverage maps, and contact forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your tech business doesn't have a website yet.\n\n"
            "Tech-savvy customers expect businesses to have an online presence.\n\n"
            "I build websites for electronics and tech businesses — with "
            "product catalogs, service menus, and location info.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
    },

    "Printing & Media": {
        "print": (
            "Hi {name} 👋\n\n"
            "I noticed your print shop doesn't have a website yet.\n\n"
            "Businesses search for printing services online. "
            "A website with your portfolio and price list "
            "wins jobs before they even call.\n\n"
            "I build websites for print shops — with service catalogs, "
            "pricing pages, sample galleries, and order forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "stationery": (
            "Hi {name} 👋\n\n"
            "I noticed your stationery shop doesn't have a website yet.\n\n"
            "Students and office managers search for stationery online.\n\n"
            "I build websites for stationery shops — with product catalogs, "
            "school/office supply bundles, and bulk order forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your media business doesn't have a website yet.\n\n"
            "A website with a portfolio can bring in clients who "
            "search online for printing/advertising services.\n\n"
            "I build professional websites for media businesses — with "
            "portfolio galleries, service lists, and pricing.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
    },

    "Supermarkets & Grocery": {
        "supermarket": (
            "Hi {name} 👋\n\n"
            "I noticed your supermarket doesn't have a website yet.\n\n"
            "Shoppers search for supermarkets and weekly deals online.\n\n"
            "I build websites for supermarkets — with weekly flyer pages, "
            "product categories, delivery info, and WhatsApp ordering.\n\n"
            "Ready in 3 weeks. Interested?"
        ),
        "grocery": (
            "Hi {name} 👋\n\n"
            "I noticed your grocery store doesn't have a website yet.\n\n"
            "People search for fresh food and grocery shops near them.\n\n"
            "I build websites for grocery stores — with product lists, "
            "daily specials, delivery zones, and WhatsApp orders.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "convenience": (
            "Hi {name} 👋\n\n"
            "I noticed your shop doesn't have a website yet.\n\n"
            "People search for nearby convenience stores, especially "
            "for late-night essentials.\n\n"
            "I build websites for convenience stores — with product highlights, "
            "operating hours, delivery options, and location map.\n\n"
            "Ready in 1 week. Interested?"
        ),
        "wholesale": (
            "Hi {name} 👋\n\n"
            "I noticed your wholesale business doesn't have a website yet.\n\n"
            "Retailers and bulk buyers search for wholesalers online.\n\n"
            "I build websites for wholesale businesses — with product catalogs, "
            "bulk pricing, MOQ info, and inquiry forms.\n\n"
            "Ready in 2 weeks. Interested?"
        ),
        "_default": (
            "Hi {name} 👋\n\n"
            "I noticed your store doesn't have a website yet.\n\n"
            "Shoppers often search online for stores near them.\n\n"
            "I build websites for stores — "
            "with product lists, deals, delivery info, and WhatsApp ordering.\n\n"
            "Ready in 2-3 weeks. Interested?"
        ),
    },
}

# ── GLOBAL FALLBACK ───────────────────────────────────────────
DEFAULT_WHATSAPP = (
    "Hi {name} 👋\n\n"
    "I noticed your business doesn't have a website yet.\n\n"
    "A professional website can help you get more customers "
    "and look credible online.\n\n"
    "I build clean, fast websites for Nigerian businesses — "
    "with your services, photos, and contact info.\n\n"
    "Ready in 2-3 weeks. Interested?"
)


def _pick_niche_template(types_str: str, category: str) -> str:
    """
    Given the OSM type string (e.g. "hairdresser", "pharmacy")
    and the category, pick the most specific template.
    """
    cat_templates = NICHE_TEMPLATES.get(category, {})
    types_lower = types_str.lower()

    # Try each keyword in the type string against the template keys
    for keyword in cat_templates:
        if keyword.startswith("_"):
            continue
        if keyword in types_lower:
            return cat_templates[keyword]

    # Category-level default
    if "_default" in cat_templates:
        return cat_templates["_default"]

    return DEFAULT_WHATSAPP


def make_whatsapp_msg(name: str, category: str, types: str = "") -> str:
    template = _pick_niche_template(types, category)
    return template.format(name=name, demo_link=DEMO_LINK)


# ── CLEAN PHONE NUMBERS ───────────────────────────────────────
def clean_phone(phone: str) -> str:
    if not isinstance(phone, str) or not phone.strip():
        return ""
    p = phone.strip().replace(" ", "").replace("-", "").replace(".", "")
    if p.startswith("0") and len(p) >= 10:
        p = "+234" + p[1:]
    if p.startswith("234") and not p.startswith("+"):
        p = "+" + p
    return p


# ── CLEAN ADDRESS ─────────────────────────────────────────────
def clean_address(addr: str) -> str:
    if not isinstance(addr, str):
        return ""
    parts = [p.strip() for p in addr.split(",")]
    seen, out = set(), []
    for p in parts:
        if p.lower() not in seen and p:
            seen.add(p.lower())
            out.append(p)
    return ", ".join(out)


# ── SCORE LABEL ───────────────────────────────────────────────
def score_label(score) -> str:
    try:
        s = int(score)
        if s >= 60:
            return "🔥 Hot"
        if s >= 35:
            return "⚡ Warm"
        return "❄️ Cold"
    except (ValueError, TypeError):
        return "❄️ Cold"


# ── MAIN ──────────────────────────────────────────────────────
def main():
    print(f"\n{Fore.YELLOW}{'━' * 52}")
    print(f"  🧹  Lead List Cleaner (Multi-City)")
    print(f"{'━' * 52}{Style.RESET_ALL}\n")

    # Load
    try:
        df = pd.read_csv(INPUT_CSV)
    except FileNotFoundError:
        print(f"{Fore.RED}❌  {INPUT_CSV} not found.")
        print(f"    Run restaurant_hunter.py first.{Style.RESET_ALL}")
        return

    print(f"{Fore.CYAN}Loaded {len(df)} rows from {INPUT_CSV}{Style.RESET_ALL}")

    # ── Clean each column ──────────────────────────────────────
    df["name"]     = df["name"].astype(str).str.strip().str.title()
    df["phone"]    = df["phone"].astype(str).apply(clean_phone)
    df["address"]  = df["address"].astype(str).apply(clean_address)
    df["city"]     = df["city"].astype(str).str.strip().str.title()
    df["category"] = df["category"].astype(str).str.strip()

    # ── Drop rows with no name ─────────────────────────────────
    df = df[df["name"].notna() & (df["name"] != "") & (df["name"] != "Nan")]

    # ── Deduplicate by name + phone + city ─────────────────────
    before = len(df)
    df = df.drop_duplicates(subset=["name", "phone", "city"], keep="first")
    print(f"{Fore.CYAN}Removed {before - len(df)} duplicates{Style.RESET_ALL}")

    # ── Add priority label ─────────────────────────────────────
    if "lead_score" in df.columns:
        df["priority"] = df["lead_score"].apply(score_label)
    else:
        df["priority"] = "❄️ Cold"

    # ── Add WhatsApp message ───────────────────────────────────
    df["whatsapp_message"] = df.apply(
        lambda row: make_whatsapp_msg(
            row["name"],
            row.get("category", ""),
            row.get("types", ""),
        ),
        axis=1,
    )

    # ── Add status column (for tracking outreach) ──────────────
    df["status"] = "Not Contacted"

    # ── Add notes column ──────────────────────────────────────
    df["notes"] = ""

    # ── Select and reorder final columns ──────────────────────
    final_cols = [
        "priority",
        "city",
        "category",
        "name",
        "phone",
        "types",
        "address",
        "rating",
        "reviews",
        "source",
        "status",
        "notes",
        "whatsapp_message",
    ]
    final_cols = [c for c in final_cols if c in df.columns]
    df = df[final_cols]

    # ── Sort: Hot first, then by rating ───────────────────────
    priority_order = {"🔥 Hot": 0, "⚡ Warm": 1, "❄️ Cold": 2}
    df["_sort"] = df["priority"].map(priority_order).fillna(3)
    df = df.sort_values(["_sort", "rating"], ascending=[True, False])
    df = df.drop(columns=["_sort"])
    df = df.reset_index(drop=True)
    df.index += 1

    # ── Save CSV ───────────────────────────────────────────────
    df.to_csv(OUTPUT_CSV, index_label="#")
    print(f"{Fore.GREEN}✅  CSV  → {OUTPUT_CSV}{Style.RESET_ALL}")

    # ── Save Excel with formatting ─────────────────────────────
    try:
        with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Leads", index_label="#")
            ws = writer.sheets["Leads"]

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            header_fill = PatternFill("solid", fgColor="1A1108")
            header_font = Font(bold=True, color="C8A96E", size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            hot_fill  = PatternFill("solid", fgColor="2D1A0E")
            warm_fill = PatternFill("solid", fgColor="1A1A0E")
            cold_fill = PatternFill("solid", fgColor="111111")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                priority_val = str(row[1].value) if len(row) > 1 else ""
                fill = (
                    hot_fill if "Hot" in priority_val else
                    warm_fill if "Warm" in priority_val else
                    cold_fill
                )
                for cell in row:
                    cell.fill = fill
                    cell.font = Font(color="FAF7F2", size=10)
                    cell.alignment = Alignment(
                        vertical="center",
                        wrap_text=(cell.column == ws.max_column),
                    )

            col_widths = {
                "priority": 10, "city": 14, "category": 18, "name": 28,
                "phone": 18, "types": 16, "address": 38, "rating": 8,
                "reviews": 10, "source": 14, "status": 16,
                "notes": 20, "whatsapp_message": 60,
            }
            for i, col_name in enumerate(df.columns, start=2):
                letter = get_column_letter(i)
                width = col_widths.get(col_name, 16)
                ws.column_dimensions[letter].width = width

            ws.row_dimensions[1].height = 22
            for i in range(2, ws.max_row + 1):
                ws.row_dimensions[i].height = 18

            ws.freeze_panes = "B2"

        print(f"{Fore.GREEN}✅  Excel → {OUTPUT_XLSX}{Style.RESET_ALL}")
    except Exception as e:
        print(f"{Fore.YELLOW}⚠️  Excel export failed: {e}{Style.RESET_ALL}")

    # ── Terminal preview (top 20) ──────────────────────────────
    preview_cols = ["priority", "city", "category", "name", "phone", "address", "status"]
    preview = df[[c for c in preview_cols if c in df.columns]].head(20)

    table = tabulate(
        preview,
        headers="keys",
        tablefmt="rounded_outline",
        showindex=True,
        maxcolwidths=[4, 4, 14, 18, 26, 18, 32, 16],
    )

    print(f"\n{Fore.YELLOW}── Top 20 Leads Preview ──────────────────────────{Style.RESET_ALL}")
    print(table)

    with open(OUTPUT_TXT, "w") as f:
        f.write(table)

    # ── Summary ────────────────────────────────────────────────
    hot  = len(df[df["priority"] == "🔥 Hot"])
    warm = len(df[df["priority"] == "⚡ Warm"])
    cold = len(df[df["priority"] == "❄️ Cold"])
    with_phone = len(df[df["phone"] != ""])

    print(f"\n{Fore.YELLOW}── Summary ───────────────────────────────────────{Style.RESET_ALL}")
    print(f"  Total leads     : {Fore.WHITE}{len(df)}{Style.RESET_ALL}")
    print(f"  🔥 Hot          : {Fore.RED}{hot}{Style.RESET_ALL}")
    print(f"  ⚡ Warm         : {Fore.YELLOW}{warm}{Style.RESET_ALL}")
    print(f"  ❄️  Cold         : {Fore.CYAN}{cold}{Style.RESET_ALL}")
    print(f"  📞 With phone   : {Fore.GREEN}{with_phone}{Style.RESET_ALL}  ← targets today")

    # City breakdown
    print(f"\n{Fore.YELLOW}── By City ──{Style.RESET_ALL}")
    for city, count in df["city"].value_counts().head(15).items():
        print(f"  {city:20s} {count}")

    # Category breakdown
    print(f"\n{Fore.YELLOW}── By Category ──{Style.RESET_ALL}")
    for cat, count in df["category"].value_counts().head(15).items():
        print(f"  {cat:20s} {count}")

    print(f"\n  Open {Fore.GREEN}{OUTPUT_XLSX}{Style.RESET_ALL} in Excel or Google Sheets\n")


if __name__ == "__main__":
    main()
